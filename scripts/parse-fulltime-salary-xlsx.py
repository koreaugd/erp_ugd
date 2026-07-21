# -*- coding: utf-8 -*-
"""
지점별 정직원 급여대장 엑셀(급여대장7월/) → ERP 급여대장 스키마 JSON.

  python scripts/parse-fulltime-salary-xlsx.py            # 대조표만 출력
  python scripts/parse-fulltime-salary-xlsx.py -o out.json # JSON 도 저장

읽기 전용이다. Firestore 에 쓰는 것은 scripts/seed-fulltime-salary.mjs 가 한다.

헤더 텍스트로 컬럼을 찾는다(고정 인덱스 금지) — 지점마다 '직급' 유무·추가 컬럼(연말보너스/이전근무지)이
달라서 인덱스로 읽으면 한 칸씩 밀린 값이 급여로 들어간다.
"""
import argparse
import io
import json
import os
import re
import sys
import warnings
from datetime import datetime, timedelta

warnings.filterwarnings("ignore")
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

XLSX_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "급여대장7월")
MONTH = "2026-07"

# ── 파일 → 지점/시트. 시트명·A1 제목이 실제 지점과 다른 파일이 있어(카츠스위스 시트명='금샤빠…',
#    대골뼈국 A1='연하동 신논현점') 여기서 명시적으로 못 박는다.
FILES = [
    ("2026.07 대학로 고래 직원 급여내역.xlsx",      "26년 7월",            "대학로고래"),
    ("2026년 07월 카라멘야 직원급여.xlsx",           "직원급여",             "카라멘야"),
    ("2607 연하동(연남점)UGD 정직원 인건비..xlsx",   "Sheet1",              "연하동 연남본점"),
    ("8번대물집 인건비 대장.xlsx",                    "2607",                "8번대물집"),
    ("단단 급여내역.xlsx",                            "직원급여",             "사카바단단"),
    ("대학로 연하동 7월직원 급여내역.xlsx",          "Sheet1",              "연하동 대학로점"),  # 한셀 파일
    ("오키스테이크하우스 7월 정직원 인건비 대장.xlsx", "Sheet1",             "오키스테이크하우스"),
    ("인건비파일_대골뼈국 (1).xlsx",                  "7월",                 "강남대골뼈국"),
    ("종로점_인건비파일 (1).xlsx",                    "26년 7월",            "대물섬 종로점"),
    ("카츠스위스 7월 정직원 인건비 대장.xlsx",        "금샤빠(압구정점)_2404", "카츠스위스"),
    ("한남동 남산광어 언건비 대장 7월.xlsx",          "남산광어",             "남산광어"),
    ("한남동 대물섬 인건비 대장 7월.xlsx",            "대물섬 한남(5월)",     "대물섬 한남점"),
]

# ── 사용자 확정 규칙(2026-07-21) ──
# 이름 교정: 엑셀 표기 → ERP 직원명부 표기. "ERP 기준으로" 라고 지시받은 건만 넣는다.
NAME_FIX = {
    "대학로고래":    {"유 현": "유현"},
    "대물섬 종로점": {"정소연": "정소영"},
}
# 소속이 다른 지점 파일에 섞여 들어온 행은 제외한다(이지현·양희지는 카츠스위스 소속으로 확정).
EXCLUDE = {
    "오키스테이크하우스": {"이지현"},
}
# 엑셀 이름 그대로 새 행으로 넣을 사람(직원명부의 비슷한 이름과 합치지 않는다).
#   카라멘야 Piao junwei ↔ 명부 박준위 = 다른 사람 / 8번대물집 박재형 ↔ 명부 박이도 = 다른 사람
FORCE_NEW_ROW = {
    "카라멘야":   {"Piao junwei"},
    "8번대물집":  {"박재형"},
}

# ── 헤더 텍스트 → ERP 필드 ──
HEADER_MAP = {
    "성명": "name",
    "직급": "rank",
    "주민등록번호": "residentNumber",
    "입사일": "entryDate",
    "근로계약": "contractType",
    "입금계좌": "_account",
    "전월급여": "prevSalary",
    "이달급여": "thisSalary",
    "이달월급": "thisSalary",
    "근무시간": "overtimeHours",
    "근무시간(직원은연장시간)": "overtimeHours",
    "시급": "overtimeRate",
    "계": "_otTotal",            # 자동계산 — 저장하지 않고 검산에만 쓴다
    "택시비및기타지출": "taxiEtc",
    "유류비및기타지출": "taxiEtc",
    "유류비": "taxiEtc",
    "상여금": "bonusTip",
    "상여금(팁)": "bonusTip",
    "총금액": "_grandTotal",     # 자동계산 — 검산용
    "실수령액(송금액)": "_netPay",  # ERP에 대응 필드 없음
    "현재근무지": "remitBranch",
    "실제송금지점": "remitBranch",
    "이전근무지": "_prevBranch",
    "연말보너스": "_yearEndBonus",
    "명절상여금": "_holidayBonus",
    "기타내용(퇴사일및퇴직금등)": "memo",
}
ERROR_TOKENS = ("#REF!", "#NAME?", "#VALUE!", "#DIV/0!", "#N/A", "#NULL!", "#NUM!")


def norm_header(v):
    return re.sub(r"\s+", "", str(v or "")).replace("\n", "")


def is_error(v):
    return isinstance(v, str) and any(tok in v for tok in ERROR_TOKENS)


def money(v):
    """ERP cleanNumeric 과 같은 규칙: 숫자만 남긴다. 0·빈값·수식에러는 빈 문자열."""
    if v is None or is_error(v):
        return ""
    if isinstance(v, (int, float)):
        return "" if int(round(v)) == 0 else str(int(round(v)))
    digits = re.sub(r"[^0-9]", "", str(v))
    if not digits:
        return ""
    return "" if int(digits) == 0 else str(int(digits))


def hours(v):
    """연장 근무시간은 소수 허용(2.5). 0·빈값·에러는 빈 문자열."""
    if v is None or is_error(v):
        return ""
    if isinstance(v, (int, float)):
        f = float(v)
        return "" if f == 0 else (str(int(f)) if f == int(f) else str(f))
    s = re.sub(r"[^0-9.]", "", str(v))
    if not s:
        return ""
    try:
        f = float(s)
    except ValueError:
        return ""
    return "" if f == 0 else (str(int(f)) if f == int(f) else str(f))


def to_date(v, warns, who):
    """입사일 → YYYY-MM-DD. 엑셀 일련번호·'2022.03.05'·datetime 모두 변환, 수식에러는 버린다."""
    if v is None or v == "":
        return ""
    if is_error(v):
        warns.append(f"{who}: 입사일이 엑셀 수식 오류({str(v).strip()}) → 빈칸으로 둠")
        return ""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, (int, float)):
        # 엑셀 일련번호(1899-12-30 기준). 1900 윤년 버그 구간(<61)은 애초에 급여 입사일로 올 수 없다.
        d = datetime(1899, 12, 30) + timedelta(days=int(v))
        return d.strftime("%Y-%m-%d")
    m = re.match(r"^\s*(\d{4})[.\-/\s]+(\d{1,2})[.\-/\s]+(\d{1,2})", str(v))
    if m:
        return f"{m.group(1)}-{m.group(2).zfill(2)}-{m.group(3).zfill(2)}"
    warns.append(f"{who}: 입사일 '{str(v).strip()}' 을 날짜로 못 읽음 → 빈칸으로 둠")
    return ""


def resident(v):
    """ERP formatResidentNumber 와 동일: 숫자 13자리 + 하이픈."""
    if v is None or is_error(v):
        return ""
    digits = re.sub(r"\D", "", str(v))[:13]
    if len(digits) <= 6:
        return digits
    return f"{digits[:6]}-{digits[6:]}"


def split_account(v):
    """'국민 000000000000' / '000000000000 국민' → (은행, 계좌). 은행명이 앞뒤 어디 있어도 된다.
    ERP splitLegacyAccount 와 같은 결과. (실계좌를 예시로 적지 않는다 — .gitignore 가 막는 항목이다)"""
    s = str(v or "").strip()
    if not s or is_error(s):
        return "", ""
    bank = re.sub(r"[0-9\-./() ]", "", s).strip()
    number = re.sub(r"[^0-9-]", "", s).strip()
    return bank, number


def text(v):
    if v is None or is_error(v):
        return ""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return re.sub(r"\s+", " ", str(v)).strip()


# ── 시트 읽기: openpyxl 이 못 여는 한셀 파일은 XML 을 직접 판다 ──
def read_grid(path, sheet_name):
    """지정한 이름의 시트를 {행번호: {열문자: 값}} 으로 돌려준다.

    시트를 '이름으로' 찾지 못하면 반드시 예외를 던진다. 조용히 다른 시트로 넘어가면 안 된다 —
    대부분의 파일이 월별 시트를 여러 장 갖고 있어(대학로고래 26년 5·6·7월, 종로점 2~7월,
    8번대물집 2605·2606·2607) 엉뚱한 시트를 읽으면 '지난달 급여가 이번 달 급여로' 조용히 들어간다.
    """
    workbook = None
    try:
        import openpyxl
        workbook = openpyxl.load_workbook(path, data_only=True)
    except Exception:
        # 워크북 자체를 못 연 경우만 한셀 폴백으로 넘어간다.
        # (여기서 시트 없음까지 삼키면 폴백이 첫 시트를 읽어버린다 — 그게 바로 위에서 말한 사고다.)
        workbook = None

    if workbook is not None:
        if sheet_name not in workbook.sheetnames:
            raise RuntimeError(
                f"{os.path.basename(path)}: '{sheet_name}' 시트가 없습니다. "
                f"이 파일의 시트: {workbook.sheetnames}"
            )
        ws = workbook[sheet_name]
        grid = {}
        for r in range(1, ws.max_row + 1):
            row = {}
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(r, c)
                if cell.value not in (None, ""):
                    row[cell.column_letter] = cell.value
            if row:
                grid[r] = row
        return grid

    # 한셀(HCell)로 저장된 파일은 x: 네임스페이스 접두사 때문에 openpyxl 이 열지 못한다.
    import zipfile
    import xml.etree.ElementTree as ET
    NS = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    REL = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"
    PKG_REL = "{http://schemas.openxmlformats.org/package/2006/relationships}"
    z = zipfile.ZipFile(path)
    strings = []
    if "xl/sharedStrings.xml" in z.namelist():
        for si in ET.fromstring(z.read("xl/sharedStrings.xml")).iter(NS + "si"):
            strings.append("".join(t.text or "" for t in si.iter(NS + "t")))

    # 여기서도 이름으로 찾는다. workbook.xml 의 <sheet name= r:id=> 를 rels 로 풀어 실제 XML 경로를 얻는다.
    sheets = {s.get("name"): s.get(REL + "id") for s in ET.fromstring(z.read("xl/workbook.xml")).iter(NS + "sheet")}
    if sheet_name not in sheets:
        raise RuntimeError(
            f"{os.path.basename(path)}: '{sheet_name}' 시트가 없습니다(한셀 파일). "
            f"이 파일의 시트: {list(sheets)}"
        )
    targets = {rel.get("Id"): rel.get("Target") for rel in ET.fromstring(z.read("xl/_rels/workbook.xml.rels")).iter(PKG_REL + "Relationship")}
    target = targets.get(sheets[sheet_name])
    if not target:
        raise RuntimeError(f"{os.path.basename(path)}: '{sheet_name}' 시트의 실제 경로를 찾지 못했습니다.")
    sheet_xml = target if target.startswith("xl/") else f"xl/{target.lstrip('/')}"
    if sheet_xml not in z.namelist():
        raise RuntimeError(f"{os.path.basename(path)}: 시트 파일 {sheet_xml} 이 없습니다.")

    grid = {}
    for row in ET.fromstring(z.read(sheet_xml)).iter(NS + "row"):
        r = int(row.get("r"))
        cells = {}
        for c in row.iter(NS + "c"):
            col = re.match(r"[A-Z]+", c.get("r")).group()
            v = c.find(NS + "v")
            val = v.text if v is not None else None
            if val is None:
                inline = c.find(NS + "is")
                if inline is not None:
                    val = "".join(x.text or "" for x in inline.iter(NS + "t"))
            elif c.get("t") == "s":
                val = strings[int(val)]
            elif val is not None:
                try:
                    f = float(val)
                    val = int(f) if f == int(f) else f
                except ValueError:
                    pass
            if val not in (None, ""):
                cells[col] = val
        if cells:
            grid[r] = cells
    return grid


def row_overtime_pay(row):
    """ERP rowOvertimePay 와 동일: 시간·시급이 '둘 다' 있을 때만 시간×시급."""
    h = float(row["overtimeHours"]) if row["overtimeHours"] else 0.0
    rate = int(row["overtimeRate"]) if row["overtimeRate"] else 0
    return round(h * rate) if (h > 0 and rate > 0) else 0


def parse_sheet(path, sheet_name, branch, warns, sources=None):
    grid = read_grid(path, sheet_name)
    # 어느 파일의 어느 시트를 읽었는지, 그 시트 A1 제목이 무엇인지 남긴다.
    # 시트가 여러 장인 파일이 많아, 사람이 눈으로 '7월 맞나'를 확인할 수 있어야 한다.
    if sources is not None:
        sources.append((branch, os.path.basename(path), sheet_name, text(grid.get(1, {}).get("A"))))

    # 헤더 행 = '성명' 이 있는 행. 그 아래 행에 '근무시간' 이 있으면 2단 헤더다.
    hdr_row = next((r for r in sorted(grid) if any(norm_header(v) == "성명" for v in grid[r].values())), None)
    if hdr_row is None:
        raise RuntimeError(f"{branch}: '성명' 헤더를 찾지 못함")
    col_of = {}
    for col, v in grid[hdr_row].items():
        key = HEADER_MAP.get(norm_header(v))
        if key:
            col_of[col] = key
    sub = grid.get(hdr_row + 1, {})
    two_row = any(norm_header(v) in ("근무시간", "근무시간(직원은연장시간)") for v in sub.values())
    if two_row:
        for col, v in sub.items():
            key = HEADER_MAP.get(norm_header(v))
            if key:
                col_of[col] = key
    data_start = hdr_row + (2 if two_row else 1)

    name_col = next((c for c, k in col_of.items() if k == "name"), "A")
    fixes = NAME_FIX.get(branch, {})
    excludes = EXCLUDE.get(branch, set())

    rows, notes = [], []
    for r in sorted(grid):
        if r < data_start:
            continue
        raw = grid[r]
        nm = text(raw.get(name_col))
        # '합계' 를 만나면 끝. 아래의 '본사 지원' 블록은 정직원 급여대장이 아니므로 읽지 않는다.
        if nm in ("합계", "합 계"):
            break
        if not nm:
            continue

        vals = {}
        extra_text = []
        for col, v in raw.items():
            key = col_of.get(col)
            if key is None:
                # 헤더 없는 칸에 적힌 메모(카츠스위스 양희지 '총 연차3일 사용' 등)는 비고에 붙인다.
                t = text(v)
                if t and col > name_col:
                    extra_text.append(t)
                continue
            vals[key] = v

        fixed = fixes.get(nm, nm)
        if fixed in excludes or nm in excludes:
            notes.append(f"제외: {nm} (다른 지점 소속으로 확정)")
            continue
        if fixed != nm:
            notes.append(f"이름 교정: '{nm}' → '{fixed}' (ERP 직원명부 기준)")

        who = f"{branch}/{fixed}"
        bank, account = split_account(vals.get("_account"))
        memo_parts = [text(vals.get("memo"))] + extra_text
        row = {
            "name": fixed,
            "rank": text(vals.get("rank")),
            "residentNumber": resident(vals.get("residentNumber")),
            "entryDate": to_date(vals.get("entryDate"), warns, who),
            "contractType": text(vals.get("contractType")) or "4대보험",
            "bank": bank,
            "accountNumber": account,
            "prevSalary": money(vals.get("prevSalary")),
            "thisSalary": money(vals.get("thisSalary")),
            "overtimeHours": hours(vals.get("overtimeHours")),
            "overtimeRate": money(vals.get("overtimeRate")),
            "overtimePay": "",
            "taxiEtc": money(vals.get("taxiEtc")),
            "bonusTip": money(vals.get("bonusTip")),
            "remitBranch": text(vals.get("remitBranch")),
            "memo": " / ".join([p for p in memo_parts if p]),
            "forceNewRow": fixed in FORCE_NEW_ROW.get(branch, set()),
            "_srcRow": r,
            "_xlGrandTotal": money(vals.get("_grandTotal")),
        }
        # 근로계약 표기 흔들림 정리('4대 보험' → '4대보험'). ERP 드롭다운 값과 맞춘다.
        row["contractType"] = re.sub(r"\s+", "", row["contractType"]) or "4대보험"

        # ── 이달급여를 엑셀 '총금액'에 맞춘다(사용자 확정 규칙, 2026-07-21) ──
        # ERP 에는 총금액 입력칸이 없다(이달급여+택시비+상여금+연장계 자동합산). 그래서 지점이 총금액 칸에
        # 직접 적어 넣은 실제 지급액(중도입사·병가·파견 일할계산)을 재현하려면 이달급여를 역산해야 한다.
        # 대부분의 지점은 두 값이 같아 역산해도 그대로다. 값이 바뀌는 행만 원래 월 정액을 비고에 남긴다.
        xl_total = row["_xlGrandTotal"]
        row["_adjusted"] = ""
        if xl_total:
            derived = int(xl_total) - int(row["taxiEtc"] or 0) - int(row["bonusTip"] or 0) - row_overtime_pay(row)
            current = int(row["thisSalary"] or 0)
            if derived < 0:
                # 역산이 음수면 엑셀 총금액이 깨진 것이다. 급여를 마이너스로 만들지 않고 원본을 지킨다.
                warns.append(f"{who}: 총금액({int(xl_total):,})이 택시비·상여금 합보다 작아 역산 불가 → 엑셀 이달급여 유지")
            elif derived != current:
                row["thisSalary"] = str(derived) if derived else ""
                row["_adjusted"] = f"{current:,} → {derived:,}"
                if current:
                    row["memo"] = (row["memo"] + " / " if row["memo"] else "") + f"월 정액 {current:,}원"

        # ERP 에 대응 필드가 없는 값이 실제로 쓰였는지 확인 — 비어 있으면 손실이 아니다.
        for k, label in (("_yearEndBonus", "연말 보너스"), ("_holidayBonus", "명절 상여금"),
                         ("_prevBranch", "이전 근무지")):
            if money(vals.get(k)) or (k == "_prevBranch" and text(vals.get(k))):
                warns.append(f"{who}: '{label}' 에 값이 있으나 ERP 에 대응 칸이 없음 → 버려짐")
        rows.append(row)
    return rows, notes


def erp_total(row):
    """ERP 가 화면에 그릴 총금액(rowTotal)."""
    n = lambda v: int(v) if str(v).isdigit() else 0
    return n(row["thisSalary"]) + n(row["taxiEtc"]) + n(row["bonusTip"]) + row_overtime_pay(row)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("-o", "--out", help="JSON 저장 경로")
    args = ap.parse_args()

    warns, result, all_notes, sources = [], {}, [], []
    for fn, sheet, branch in FILES:
        path = os.path.join(XLSX_DIR, fn)
        rows, notes = parse_sheet(path, sheet, branch, warns, sources)
        result[branch] = rows
        all_notes += [f"{branch}: {n}" for n in notes]

    print(f"# 정직원 급여대장 {MONTH} — 엑셀 파싱 결과 (읽기 전용)\n")
    print("## 읽은 시트 (파일에 시트가 여러 장이라 '이 달이 맞는지' 눈으로 확인하세요)")
    for branch, fn, sheet, title in sources:
        print(f"  {branch:<14} {fn}  →  시트 '{sheet}'  (A1: {title or '-'})")
    print()
    grand, mismatch, adjusted = 0, [], []
    for branch, rows in result.items():
        total = sum(erp_total(r) for r in rows)
        grand += total
        print(f"## {branch} — {len(rows)}명 / ERP 총액 {total:,}원")
        print(f"{'성명':<12}{'직급':<6}{'입사일':<12}{'은행':<8}{'전월':>10}{'이달':>10}"
              f"{'연장h':>7}{'시급':>8}{'택시등':>9}{'상여':>9}{'ERP총액':>11}{'엑셀총액':>11}  비고")
        for r in rows:
            e, x = erp_total(r), int(r["_xlGrandTotal"] or 0)
            flag = ""
            if x and e != x:
                flag = "  ⚠엑셀과 다름"
                mismatch.append(f"{branch}/{r['name']}: ERP {e:,} vs 엑셀 {x:,}")
            if r["_adjusted"]:
                flag += f"  ← 이달급여 {r['_adjusted']}"
                adjusted.append(f"{branch}/{r['name']}: 이달급여 {r['_adjusted']}")
            nm = r["name"] + ("*" if r["forceNewRow"] else "")
            print(f"{nm:<12}{r['rank'] or '-':<6}{r['entryDate'] or '-':<12}{r['bank'] or '-':<8}"
                  f"{int(r['prevSalary'] or 0):>10,}{int(r['thisSalary'] or 0):>10,}"
                  f"{r['overtimeHours'] or '-':>7}{int(r['overtimeRate'] or 0):>8,}"
                  f"{int(r['taxiEtc'] or 0):>9,}{int(r['bonusTip'] or 0):>9,}"
                  f"{e:>11,}{x:>11,}{flag}")
            if r["memo"]:
                print(f"    └ {r['memo'][:150]}")
        print()

    print(f"## 합계: {sum(len(v) for v in result.values())}명 / {grand:,}원\n")
    if adjusted:
        print(f"## 엑셀 총금액에 맞춰 이달급여를 역산한 행 ({len(adjusted)}건)")
        for a in adjusted:
            print("  -", a)
        print()
    print(f"## 엑셀 총금액과 ERP 자동계산 불일치: {len(mismatch)}건")
    for m in mismatch:
        print("  -", m)
    print()
    if all_notes:
        print("## 적용한 이름 규칙")
        for n in all_notes:
            print("  -", n)
        print()
    if warns:
        print("## 확인 필요")
        for w in warns:
            print("  -", w)
        print()

    if args.out:
        payload = {"month": MONTH, "branches": result}
        with open(args.out, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)
        print(f"JSON 저장: {args.out}")


if __name__ == "__main__":
    main()
