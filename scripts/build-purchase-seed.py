"""거래처 이체리스트 엑셀 → ERP 매입매출(monthly_purchases) 시드 JSON 변환.

일회성 스크립트. 엑셀을 읽어 ERP의 PurchaseSalesRow 형식으로 바꾸고,
지점별 행수/합계가 엑셀과 일치하는지 검증한 뒤 JSON으로 내보낸다.
실제 Firestore 쓰기는 seed-monthly-purchases.mjs 가 담당한다.

사용법:
  python scripts/build-purchase-seed.py <엑셀경로> <대상월 YYYY-MM> <출력 JSON경로>
"""
import json
import re
import sys

import openpyxl

# 엑셀 시트명 → ERP public_branches 의 branchName. None 이면 건너뛴다.
BRANCH_MAP = {
    "대물섬종로점": "대물섬 종로점",
    "연하동연남점": "연하동 연남본점",
    "사카바단단": "사카바단단",
    "대골뼈국": "강남대골뼈국",
    "마음죽": "마음죽",
    "BBQ험프리스점": None,  # ERP에 지점 없음 — 이번 시드에서 제외
    "남산광어": "남산광어",
    "대학로 고래": "대학로고래",
    "연하동 대학로": "연하동 대학로점",
    "오키스테이크 하우스": "오키스테이크하우스",
    "카츠스위스": "카츠스위스",
    "카라멘야": "카라멘야",
    "대물섬 한남점": "대물섬 한남점",
}

# ERP 드롭다운 값과 정확히 일치해야 한다 (MonthlyPurchaseSalesSubTab.tsx 의 category union).
CATEGORY_MAP = {
    "식재료비": "식재료비",
    "주류비": "주류비",
    "식음료외기타": "식음료외 기타",
}


def clean_memo(value):
    """ERP memo는 한 줄 input이므로 줄바꿈을 ' / '로 잇는다."""
    if value is None:
        return ""
    text = re.sub(r"[\r\n]+", " / ", str(value).strip())
    return re.sub(r"\s{2,}", " ", text)


def build(xlsx_path, month):
    workbook = openpyxl.load_workbook(xlsx_path, data_only=True)
    summary_sheet = workbook["전체요약"]
    excel_summary = {
        row[0]: (row[1], row[2])
        for row in summary_sheet.iter_rows(min_row=3, values_only=True)
        if row[0] and row[0] != "총계"
    }

    documents = []
    problems = []
    skipped = []

    for sheet_name in workbook.sheetnames[1:]:
        if sheet_name not in BRANCH_MAP:
            problems.append(f"매핑 없는 시트: {sheet_name}")
            continue
        branch_name = BRANCH_MAP[sheet_name]
        sheet = workbook[sheet_name]
        rows = []
        for source in sheet.iter_rows(min_row=3, values_only=True):
            category, vendor, amount = source[0], source[1], source[2]
            if not vendor or str(vendor).strip() == "합계":
                continue
            if category not in CATEGORY_MAP:
                problems.append(f"{sheet_name}: 알 수 없는 분류항목 {category!r} ({vendor})")
                continue
            if not isinstance(amount, (int, float)) or amount <= 0:
                problems.append(f"{sheet_name}: 금액 이상 {amount!r} ({vendor})")
                continue
            transfer = str(int(amount))
            rows.append({
                "id": f"p_{month}_{len(rows) + 1}",
                "category": CATEGORY_MAP[category],
                "vendorName": str(vendor).strip(),
                "transferAmount": transfer,
                "bank": str(source[3]).strip() if source[3] else "",
                "accountNumber": str(source[4]).strip() if source[4] else "",
                "isPrepaid": False,
                "prepaidChargeAmount": "",
                # 비선입금 행은 앱 규칙상 실제 이달사용액 = 이체필요금액.
                "monthlyUsageAmount": transfer,
                "transferNeeded": True,
                "memo": clean_memo(source[5]),
            })

        expected_count, expected_total = excel_summary.get(sheet_name, (None, None))
        actual_total = sum(int(r["transferAmount"]) for r in rows)
        if expected_count != len(rows) or expected_total != actual_total:
            problems.append(
                f"{sheet_name}: 엑셀요약({expected_count}건/{expected_total:,}) != 변환({len(rows)}건/{actual_total:,})"
            )

        if branch_name is None:
            skipped.append({"sheet": sheet_name, "rows": len(rows), "total": actual_total})
            continue

        documents.append({
            "sheet": sheet_name,
            "branchName": branch_name,
            "dataKey": f"monthly_purchases:{branch_name}:{month}",
            "rows": rows,
            "total": actual_total,
        })

    return documents, skipped, problems


def main():
    xlsx_path, month, out_path = sys.argv[1], sys.argv[2], sys.argv[3]
    documents, skipped, problems = build(xlsx_path, month)

    print(f"{'엑셀 시트':<16} {'→ ERP 지점':<16} {'행':>4} {'합계':>16}")
    print("-" * 58)
    for entry in documents:
        print(f"{entry['sheet']:<16} {entry['branchName']:<16} {len(entry['rows']):>4} {entry['total']:>16,}")
    for entry in skipped:
        print(f"{entry['sheet']:<16} {'(건너뜀)':<16} {entry['rows']:>4} {entry['total']:>16,}")

    total_rows = sum(len(e["rows"]) for e in documents)
    total_amount = sum(e["total"] for e in documents)
    print("-" * 58)
    print(f"{'쓰기 대상':<33} {total_rows:>4} {total_amount:>16,}")

    if problems:
        print("\n[문제 발견]")
        for problem in problems:
            print(f"  - {problem}")
        sys.exit(1)

    with open(out_path, "w", encoding="utf-8") as handle:
        json.dump({"month": month, "documents": documents}, handle, ensure_ascii=False, indent=2)
    print(f"\n검증 통과. 시드 JSON 저장: {out_path}")


if __name__ == "__main__":
    main()
