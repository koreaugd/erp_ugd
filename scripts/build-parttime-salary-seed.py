# -*- coding: utf-8 -*-
"""지점별 월말정산 엑셀(파트타이머급여 시트) → ERP part_time_salaries 시드 JSON.

사용법:
  python scripts/build-parttime-salary-seed.py <엑셀폴더> <대상월> <ERP급여덤프> <명부덤프> <실시간시간덤프> <출력.json> [지점 ...]

MonthlyPartTimeSalarySubTab 의 조립 규칙(E. Assemble all pieces)에 맞춰 만든다.
  · employeeId 는 반드시 직원명부의 사원 id 여야 한다 — 다르면 화면이 같은 사람 행을 하나 더 만들어
    한 사람이 두 줄이 되고 급여가 두 번 나간다. 명부에 없는 사람만 'manual-' 행으로 만든다.
  · 주민번호·입사일·은행·계좌번호·시급·팁은 저장본이 우선하므로 그대로 채우면 화면에 뜬다.
  · 누적시간은 hoursOverridden 이 켜져 있을 때만 저장본 값이 쓰인다(아니면 일일마감 집계로 덮인다).
    → 엑셀 시간을 넣으려면 켜야 하고, 켜면 그 뒤 일일마감이 갱신돼도 반영되지 않는다.
  · 급여는 화면이 늘 다시 계산한다(시급 × 시간 + 팁). 여기서도 같은 식으로 계산해 미리 채운다.
  · 출근날짜는 건드리지 않는다 — ERP는 7개까지만 저장하는데 엑셀엔 20개짜리가 있어 왕복이 안 된다.
"""
import json, os, re, sys, datetime
import openpyxl

norm = lambda s: re.sub(r"\s+", "", str(s or "").strip())
digits = lambda s: re.sub(r"\D", "", str(s or ""))


def numf(v):
    if v is None: return None
    if isinstance(v, (int, float)): return float(v)
    s = str(v).strip().replace(",", "")
    return float(s) if re.fullmatch(r"\d+(\.\d+)?", s) else None


def as_int_str(v):
    f = numf(v)
    if f is None: return ""
    return str(int(f)) if f == int(f) else str(f)


def as_date(v):
    if v is None: return ""
    if isinstance(v, (datetime.datetime, datetime.date)): return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    m = re.search(r"(\d{4})[.\-/년\s]+(\d{1,2})[.\-/월\s]+(\d{1,2})", s)
    if m: return f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
    if re.fullmatch(r"\d{8}", s): return f"{s[:4]}-{s[4:6]}-{s[6:]}"
    if re.fullmatch(r"\d{5}", s):
        return (datetime.date(1899, 12, 30) + datetime.timedelta(days=int(s))).strftime("%Y-%m-%d")
    return s


BANK_WORDS = ["국민", "신한", "우리", "하나", "농협", "기업", "카카오", "토스", "부산", "경남", "새마을", "우체국", "SC제일", "한국외환", "한국외한", "씨티"]


def split_bank_account(bank_cell, acct_cell):
    """계좌 칸에 은행명이 섞여 오는 경우가 많다('농협352-179412-9643'). 은행과 번호를 갈라 준다."""
    bank = re.sub(r"은행$", "", str(bank_cell or "").strip()).strip()
    acct = str(acct_cell or "").strip()
    if numf(acct) is not None:                      # 엑셀이 숫자로 저장한 계좌 → 소수점 제거
        acct = as_int_str(acct)
    for w in sorted(BANK_WORDS, key=len, reverse=True):
        if acct.startswith(w):
            if not bank: bank = w
            acct = acct[len(w):].strip()
            break
    acct = re.sub(r"^(은행|뱅크)\s*", "", acct).strip()
    return bank, acct


def read_excel(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    if "파트타이머급여" not in wb.sheetnames: return []
    grid = [list(r) for r in wb["파트타이머급여"].iter_rows(min_row=1, max_col=14, values_only=True)]
    out = []
    for r in grid[2:]:
        r = list(r) + [None] * (14 - len(r))
        name = str(r[0]).strip() if r[0] else ""
        if not name or name == "합계": continue
        bank, acct = split_bank_account(r[4], r[5])
        out.append({
            "name": name,
            "residentNumber": re.sub(r"\s+", "", str(r[1] or "").strip()),
            "entryDate": as_date(r[2]),
            "contractType": str(r[3] or "").strip(),
            "bank": bank, "accountNumber": acct,
            "hourlyRate": as_int_str(r[6]),
            "hours": numf(r[7]),
            "payRaw": r[8],
            "tips": as_int_str(r[10]),
            "memo": re.sub(r"\s+", " ", str(r[13] or "").strip()),
        })
    return out


def read_ods(path):
    import pandas as pd
    df = pd.read_excel(path, sheet_name="파트타이머급여", engine="odf", header=None)
    grid = [[None if pd.isna(v) else v for v in df.iloc[i, :14]] for i in range(len(df))]
    out = []
    for r in grid[2:]:
        r = list(r) + [None] * (14 - len(r))
        name = str(r[0]).strip() if r[0] else ""
        if not name or name == "합계": continue
        bank, acct = split_bank_account(r[4], r[5])
        out.append({"name": name, "residentNumber": re.sub(r"\s+", "", str(r[1] or "").strip()),
                    "entryDate": as_date(r[2]), "contractType": str(r[3] or "").strip(),
                    "bank": bank, "accountNumber": acct, "hourlyRate": as_int_str(r[6]),
                    "hours": numf(r[7]), "payRaw": r[8], "tips": as_int_str(r[10]),
                    "memo": re.sub(r"\s+", " ", str(r[13] or "").strip())})
    return out


def salary_of(rate, hours, tips):
    return str(int((float(rate or 0) * float(hours or 0)) + float(tips or 0)))


def main():
    args = sys.argv[1:]
    # --alias "지점:엑셀이름=ERP이름" — 엑셀 오타로 같은 사람이 두 줄이 되는 것을 막는다.
    aliases = {}
    rest = []
    i = 0
    while i < len(args):
        if args[i] == "--alias":
            branch, _, pair = args[i + 1].partition(":")
            xname, _, ename = pair.partition("=")
            aliases.setdefault(branch.strip(), {})[norm(xname)] = norm(ename)
            i += 2
            continue
        rest.append(args[i]); i += 1

    src, month, erp_dump, roster_dump, hours_dump, out_path = rest[:6]
    targets = rest[6:]
    FILE_BRANCH = json.load(open(os.path.join(os.path.dirname(out_path), "file_branch.json"), encoding="utf-8"))

    erp = json.load(open(erp_dump, encoding="utf-8"))
    rosters = json.load(open(roster_dump, encoding="utf-8"))["own"]
    live = json.load(open(hours_dump, encoding="utf-8"))["byBranch"]

    documents, notes, dup_warnings = [], [], []
    for base, branch in sorted(FILE_BRANCH.items(), key=lambda kv: kv[1]):
        if targets and branch not in targets: continue
        path = os.path.join(src, base)
        if not os.path.exists(path): continue
        xs = read_ods(path) if path.lower().endswith(".ods") else read_excel(path)
        xs = [x for x in xs if x["name"] not in ("해당없음",)]
        if not xs: continue

        existing = erp["salaries"].get(branch) or []
        by_name = {}
        for e in existing:
            if norm(e.get("name")): by_name.setdefault(norm(e.get("name")), []).append(e)
        roster_pt = [r for r in (rosters.get(branch) or []) if r.get("division") == "파트타이머"]
        roster_by_name = {}
        for r in roster_pt:
            if norm(r.get("name")): roster_by_name.setdefault(norm(r.get("name")), []).append(r)
        live_b = live.get(branch, {})

        alias_b = aliases.get(branch, {})
        rows, used_ids, manual_n = [], set(), 0
        for i, x in enumerate(xs):
            k = alias_b.get(norm(x["name"]), norm(x["name"]))
            if k != norm(x["name"]):
                notes.append(f"{branch} · {x['name']}: --alias 지정으로 ERP '{k}' 행에 반영 (같은 사람)")
            base_row, source = None, None
            if k in by_name and by_name[k]:
                base_row = by_name[k].pop(0); source = "erp"
            elif k in roster_by_name and roster_by_name[k]:
                base_row = {"employeeId": roster_by_name[k].pop(0)["id"]}; source = "roster"
            else:
                base_row = {"employeeId": f"manual-{month}-{branch}-{i}"}; source = "manual"
                manual_n += 1
                notes.append(f"{branch} · {x['name']}: 명부에 없어 수기 행으로 만듦")
            eid = base_row["employeeId"]
            if eid in used_ids:
                eid = f"manual-{month}-{branch}-{i}"; source = "manual"; manual_n += 1
            used_ids.add(eid)

            prev = base_row if source == "erp" else {}
            # 계좌번호: 엑셀이 숫자로 저장돼 앞의 0이 떨어진 경우가 있다. ERP 값이 0만 더 붙은 같은 번호면 ERP를 지킨다.
            acct = x["accountNumber"]
            pa = str(prev.get("accountNumber") or "")
            if pa and digits(pa).lstrip("0") == digits(acct).lstrip("0") and len(digits(pa)) >= len(digits(acct)):
                acct = pa
            # 주민번호: 엑셀이 13자리면 엑셀 우선(ERP는 앞 6자리만인 경우가 많다)
            rrn = x["residentNumber"] or str(prev.get("residentNumber") or "")
            hours_x = x["hours"]
            auto = str(live_b.get(x["name"], {}).get("hours", "")) if x["name"] in live_b else str(prev.get("autoAccumulatedHours") or "")
            if hours_x is None:
                hours_val = str(prev.get("accumulatedHours") or "")
                overridden = bool(prev.get("hoursOverridden"))
            else:
                hours_val = f"{hours_x:g}"
                overridden = True   # 엑셀 값을 지키려면 반드시 켜야 한다
            rate = x["hourlyRate"] or str(prev.get("hourlyRate") or "")
            tips = x["tips"] or str(prev.get("tipsEtcAmount") or "0") or "0"

            # 엑셀 급여칸이 숫자면 그 금액이 정답이다. ERP는 급여를 직접 못 넣고 시급×시간+팁 으로만 만들 수 있으므로,
            # 시급×시간 으로 안 채워지는 차액을 팁/기타에 넣어 총액을 맞춘다.
            #   · 시급 없이 급여만 적힌 파출·일당 → 전액이 팁/기타로 들어가고 누적시간은 엑셀 그대로 보존된다.
            #   · 시급×시간 이 급여와 딱 맞으면 차액 0 이라 팁은 건드리지 않는다.
            excel_pay = numf(x["payRaw"])
            if excel_pay is not None:
                base = (numf(rate) or 0) * (numf(hours_val) or 0)
                diff = excel_pay - base
                if abs(diff) >= 1:
                    tips = str(int(round(diff)))
                    kind = "시급 없는 파출·일당" if not rate else "시급×시간과 급여가 다름"
                    notes.append(f"{branch} · {x['name']}: {kind} — 엑셀 급여 {int(excel_pay):,} = 시급 {rate or 0}×{hours_val or 0} + 팁/기타 {int(round(diff)):,}")
                    if diff < 0:
                        dup_warnings.append(f"{branch} · {x['name']}: 팁/기타가 음수({int(diff):,}) — 시급·시간이 엑셀 급여보다 큽니다. 확인 필요")

            # 화면은 명부 이름(pt.name)을 따라가므로, 명부/기존 행에 붙인 경우 그 이름을 쓴다.
            # 엑셀 이름을 넣어 봐야 조립 한 번에 명부 이름으로 되돌아가 시드와 화면이 달라 보인다.
            display_name = str(prev.get("name") or "").strip() or x["name"] if source == "erp" else x["name"]
            row = {
                "employeeId": eid,
                "name": display_name,
                "residentNumber": rrn,
                "entryDate": x["entryDate"] or str(prev.get("entryDate") or ""),
                "contractStatus": prev.get("contractStatus") or "미작성",
                "bank": x["bank"] or str(prev.get("bank") or ""),
                "accountNumber": acct,
                "hourlyRate": rate,
                "accumulatedHours": hours_val,
                "hoursOverridden": overridden,
                "autoAccumulatedHours": auto,
                "tipsEtcAmount": tips,
                "calculatedSalary": salary_of(rate, hours_val, tips),
                "attendanceDates": str(prev.get("attendanceDates") or ""),
                "actualPaidAmount": "" if not str(prev.get("actualPaidAmount") or "") else salary_of(rate, hours_val, tips),
                "payoutBranch": str(prev.get("payoutBranch") or branch),
                "memo": x["memo"] or str(prev.get("memo") or ""),
                "edited": True,
            }
            if prev.get("rosterName"): row["rosterName"] = prev["rosterName"]
            if source == "roster":
                row["rosterName"] = x["name"]
            # 엑셀 급여칸이 '지급완료' 같은 글이면(파출·일당) 그 사실을 비고에 남긴다.
            # 이 행들은 엑셀 급여 합계에도 들어가 있지 않다 — 별도로 이미 지급된 건이다.
            if x["payRaw"] is not None and numf(x["payRaw"]) is None:
                tag = f"엑셀 급여칸: {str(x['payRaw']).strip()}"
                row["memo"] = f"{row['memo']} / {tag}".strip(" /") if row["memo"] else tag
                notes.append(f"{branch} · {x['name']}: 엑셀 급여칸이 '{str(x['payRaw']).strip()}' (숫자 아님) — 시급 {rate}, 시간 {hours_val or '없음'} → 급여 {row['calculatedSalary']}원, 비고에 기록")
            rows.append(row)

        # 엑셀에 없는 기존 행은 그대로 둔다(지우지 않는다).
        kept = [e for e in existing if e.get("employeeId") not in used_ids]
        for e in kept:
            notes.append(f"{branch} · {e.get('name')}: 엑셀에 없어 기존 행 그대로 유지 (시간 {e.get('accumulatedHours')}, 급여 {e.get('calculatedSalary')})")

        # 중복 위험: 새로 만든 수기 행과 남겨 둔 기존 행의 시급·시간이 같으면 이름만 다른 같은 사람일 수 있다.
        # 그대로 두면 한 사람에게 급여가 두 번 나간다 — 반드시 사람이 확인해야 한다.
        for m in [r for r in rows if str(r["employeeId"]).startswith("manual-")]:
            for e in kept:
                same_rate = digits(m["hourlyRate"]) and digits(m["hourlyRate"]) == digits(e.get("hourlyRate"))
                same_hours = numf(m["accumulatedHours"]) is not None and numf(m["accumulatedHours"]) == numf(e.get("accumulatedHours"))
                if same_rate and same_hours:
                    dup_warnings.append(
                        f"{branch}: 엑셀 '{m['name']}' 과 ERP '{e.get('name')}' 의 시급·시간이 같습니다 "
                        f"(시급 {m['hourlyRate']}, 시간 {m['accumulatedHours']}) — 같은 사람이면 급여가 두 번 나갑니다. "
                        f"--alias \"{branch}:{m['name']}={e.get('name')}\" 로 합치세요."
                    )
        rows.extend(kept)

        documents.append({
            "branchName": branch, "dataKey": f"part_time_salaries:{branch}:{month}",
            "rows": rows, "excelCount": len(xs), "erpCountBefore": len(existing),
            "manualCount": manual_n,
            "salaryTotal": sum(int(r["calculatedSalary"] or 0) for r in rows),
        })

    print(f"{'지점':<16}{'엑셀':>5}{'기존':>5}{'→행':>5}{'수기':>5}{'급여합계':>13}")
    print("-" * 52)
    for d in documents:
        print(f"{d['branchName']:<16}{d['excelCount']:>5}{d['erpCountBefore']:>5}{len(d['rows']):>5}{d['manualCount']:>5}{d['salaryTotal']:>13,}")
    print("-" * 52)
    print(f"{'합계':<16}{sum(d['excelCount'] for d in documents):>5}{'':>5}{sum(len(d['rows']) for d in documents):>5}"
          f"{sum(d['manualCount'] for d in documents):>5}{sum(d['salaryTotal'] for d in documents):>13,}")

    if dup_warnings:
        print(f"\n[!! 중복 위험 {len(dup_warnings)}건 — 급여 이중지급 가능]")
        for w in dup_warnings: print(f"  !! {w}")

    if notes:
        print(f"\n[검토할 것 {len(notes)}건]")
        for n in notes: print(f"  · {n}")

    json.dump({"month": month, "documents": documents}, open(out_path, "w", encoding="utf-8"), ensure_ascii=False, indent=2)
    print(f"\n시드 저장: {out_path}")
    if dup_warnings:
        print("중복 위험이 남아 있습니다. --alias 로 합치거나, 다른 사람이 맞는지 확인한 뒤 쓰세요.")
        sys.exit(2)


main()
